import os
import re
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

#=========================
# 設定
#=========================

ROOT = r"G:\old_HDD\film"
API_KEY = "af246bcc8feca07801bb5c59cd22694a"

VIDEO_EXT = {
    ".mkv",".mp4",".avi",".mov",".wmv",".iso",
    ".m2ts",".ts",".mpg",".mpeg",".flv"
}

EXCEL_COLUMN_WIDTHS = {
    "原始名稱": 65,
    "英文名稱": 35,
    "年份": 10,
    "台灣名稱": 25,
    "導演中文": 22,
    "導演英文": 22
}

# 擴充了常見的壓製組與版本字眼 (包含您列表中的常見群組)
REMOVE_WORDS = [
    "1080p","720p","2160p","480p","4k","8k",
    "BluRay","Blu-Ray","WEBRip","WEB-DL","BDRip","DVDrip",
    "HDRip","BRRip","DVDRip","Remux","AMZN",
    "H264","H265","x264","x265","HEVC","AVC",
    "AAC","AC3","DTS","DDP","DDP5","Atmos","Opus",
    "TRUEHD","HDMA","6CH", "7 1", "5 1",
    "10bit","8bit","SDR","HDR","DV",
    "ITA","ENG","JPN","CHS","CHT","GER","CHINESE","JAPANESE",
    "MULTI","SUB","SUBS","Dual","Audio",
    "Dir","Cut","Extended","Remastered","Unrated","Alternate",
    "PROPER","IMAX",
    "YTS","MX","LT","AG","AM", "RARBG", "EVO", "HazMatt", "HDiY",
    "TGx", "GalaxyRG265", "FGT", "PSA", "aXXo", "WADU", "Grym",
    "BTNET", "CnSCG", "EDGE2020", "SWAXXON", "MIRCrew", "Licdom",
    "HighCode", "DEFiNiTE", "YG", "iSCG", "heTOrico", "SWTYBLZ",
    "jeddak", "CMCT", "3Li", "Garshasp", "SUJAIDR", "PANAM"
]

#=========================
# 掃描 (修復資料夾與檔案的切割問題)
#=========================

items = []
if os.path.exists(ROOT):
    for obj in os.scandir(ROOT):
        if obj.is_file():
            base, ext = os.path.splitext(obj.name)
            if ext.lower() in VIDEO_EXT:
                # 檔案：保留原始檔名顯示，但傳入無副檔名的 base 進行清理
                items.append((obj.name, base))
        elif obj.is_dir():
            # 資料夾：直接傳入全名，不要用 splitext 切割
            items.append((obj.name, obj.name))

print("="*60)
print(f"掃描到 {len(items)} 個項目")
print("="*60)

#=========================
# 清理片名 (核心邏輯重構)
#=========================
def clean_name(clean_target):
    # 1. 先抓年份 (從字串末端往前找最合理的年份)
    # 條件：確保年份前後是邊界或特殊符號，避免抓到解析度等其他數字
    year = None
    matches = list(re.finditer(r"(?:^|[.\s\(\[_-])(19\d{2}|20\d{2})(?:[.\s\)\]_-]|$)", clean_target))
    
    title_before_year = clean_target
    if matches:
        last_match = matches[-1]
        year = last_match.group(1)
        idx = last_match.start()
        # 如果年份前面有分隔符號，從該符號處截斷
        if idx > 0:
            title_before_year = clean_target[:idx]

    # 2. 清理函式
    def purify(text):
        t = text
        # 將點和底線替換為空白
        t = t.replace(".", " ").replace("_", " ")
        # 移除括號內容
        t = re.sub(r"\[.*?\]|\(.*?\)", " ", t)
        
        # 移除中文影音站常見的後綴垃圾字眼 (改良正則，只要遇到這些字眼，後面的全砍)
        chinese_junk = r"(?:1080p|720p|4k|超清|高清|免費)?(?:在線|線上看|正片|優酷|劇迷|首選|中文字幕|簡體中字|繁中|簡中|粵語|英語).*"
        t = re.sub(chinese_junk, "", t, flags=re.I)
        
        # 移除壓製組、畫質等關鍵字
        for word in REMOVE_WORDS:
            t = re.sub(r"\b" + re.escape(word) + r"\b", " ", t, flags=re.I)
            
        # 移除非英數字與非中文字元
        t = re.sub(r"[^\w\s\u4e00-\u9fa5]", " ", t)
        # 壓縮多餘空白
        return re.sub(r"\s+", " ", t).strip()

    # title_without_year 是主要搜尋依據 (例如: Taxi)
    title_without_year = purify(title_before_year)
    # full_title 用來對付片名本身包含年份的情況 (例如: Blade Runner 2049)
    full_title = purify(clean_target)

    return title_without_year, year, full_title

#=========================
# TMDB API 操作 (多重策略)
#=========================

def search_movie(title_without_year, year, full_title):
    url = "https://api.themoviedb.org/3/search/movie"

    def do_search(q, y=None):
        if not q: return None
        params = {
            "api_key": API_KEY,
            "query": q,
            "language": "zh-TW",
            "page": 1
        }
        if y:
            params["primary_release_year"] = y
        
        try:
            r = requests.get(url, params=params, timeout=10)
            if r.status_code == 200:
                js = r.json()
                if js.get("results"):
                    return js["results"][0]["id"]
        except requests.RequestException:
            pass
        return None

    # 策略 A：去除年份的片名 + 精確年份 (對付多數常規電影)
    if year and title_without_year:
        m_id = do_search(title_without_year, year)
        if m_id: return m_id

    # 策略 B：完整片名，不限定年份 (對付 Blade Runner 2049)
    if full_title:
        m_id = do_search(full_title)
        if m_id: return m_id

    # 策略 C：放寬條件，僅搜片名無年份 (對付資料夾標錯年份的狀況)
    if title_without_year:
        m_id = do_search(title_without_year)
        if m_id: return m_id

    return None

def get_movie_details(movie_id):
    url = f"https://api.themoviedb.org/3/movie/{movie_id}"
    params = {
        "api_key": API_KEY,
        "language": "zh-TW",
        "append_to_response": "credits"
    }

    r = requests.get(url, params=params)
    if r.status_code != 200:
        return None

    js = r.json()
    directors_tw = []
    directors_en = []
    crew = js.get("credits", {}).get("crew", [])
    
    for member in crew:
        if member.get("job") == "Director":
            directors_tw.append(member.get("name", ""))
            directors_en.append(member.get("original_name", ""))
            
    return {
        "tw": js.get("title", ""),
        "original": js.get("original_title", ""),
        "director_tw": ", ".join(directors_tw),
        "director_en": ", ".join(directors_en)
    }

#=========================
# 搜尋資料
#=========================

results = []
print()
print("="*60)
print("開始搜尋電影")
print("="*60)

for raw_name, clean_target in items:
    title_without_year, year, full_title = clean_name(clean_target)

    print()
    print("原始：", raw_name)
    
    # 呈現清理後的字串供除錯參考
    display_title = title_without_year if title_without_year else full_title
    print("解析：", display_title, f"(年份: {year})" if year else "")

    movie_id = search_movie(title_without_year, year, full_title)

    if movie_id is None:
        print("找不到")
        results.append([raw_name, display_title, year, "", "", ""])
        continue

    info = get_movie_details(movie_id)
    print("英文：", info["original"])
    print("台灣：", info["tw"])
    print("導演(中)：", info["director_tw"])
    print("導演(英)：", info["director_en"])

    results.append([
        raw_name,
        info["original"],
        year,
        info["tw"],
        info["director_tw"],
        info["director_en"]
    ])

#=========================
# 匯出 Excel (XLSX)
#=========================

excel_name = "movie_list.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "電影清單"

headers = ["原始名稱", "英文名稱", "年份", "台灣名稱", "導演中文", "導演英文"]
ws.append(headers)

for row in results:
    ws.append(row)

header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
alt_fill = PatternFill(start_color="F2F6F9", end_color="F2F6F9", fill_type="solid")    
data_font = Font(name="微軟正黑體", size=10)
thin_border = Side(style='thin', color='D9D9D9')
cell_border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    
    col_letter = get_column_letter(col_idx)
    ws.column_dimensions[col_letter].width = EXCEL_COLUMN_WIDTHS.get(header, 20)

for row_idx in range(2, ws.max_row + 1):
    is_even = (row_idx % 2 == 0)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = data_font
        cell.border = cell_border
        
        if is_even:
            cell.fill = alt_fill
            
        header_name = headers[col_idx - 1]
        if header_name in ["年份"]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

wb.save(excel_name)

print()
print("="*60)
print("完成")
print("輸出:", excel_name)